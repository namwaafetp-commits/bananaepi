from __future__ import annotations

import io
from fastapi import APIRouter
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

router = APIRouter(prefix="/api/template", tags=["Template"])

# ── Color palette (dark theme matching the app) ───────────────────────────────
_HEADER_BG   = "1e2a3a"   # navy
_HEADER_FONT = "5eead4"   # teal-300
_REQ_BG      = "0f2027"   # darker for required cols
_REQ_FONT    = "f87171"   # red-400 for required label
_SEC_BG      = "162032"   # section label rows
_SEC_FONT    = "94a3b8"   # slate-400
_BOOL_BG     = "0f2027"
_BOOL_FONT   = "a3e635"   # lime for boolean cols
_BORDER_C    = "2d3f50"

_thin = Side(style="thin", color=_BORDER_C)
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

# ── Column definitions ─────────────────────────────────────────────────────────
# (name, section, type_label, accepted_values_hint, required)
_COLS: list[tuple[str, str, str, str, bool]] = [
    # Identifier
    ("case_id",               "Identifier",   "string",  "e.g. PT-001, HN123456 — unique per row",       True),

    # Person
    ("age",                   "Person",       "integer", "Whole years: 25, 3, 70",                        True),
    ("sex",                   "Person",       "string",  "Male / Female / Unknown",                       True),
    ("occupation",            "Person",       "string",  "Free text: Farmer, Student, Healthcare",        False),
    ("nationality",           "Person",       "string",  "Free text: Thai, Cambodian",                    False),
    ("underlying_ht",         "Person",       "0 / 1",   "1 = yes, 0 = no",                              False),
    ("underlying_dm",         "Person",       "0 / 1",   "1 = yes, 0 = no",                              False),
    ("underlying_copd",       "Person",       "0 / 1",   "1 = yes, 0 = no",                              False),
    ("underlying_cad",        "Person",       "0 / 1",   "1 = yes, 0 = no",                              False),
    ("underlying_ckd",        "Person",       "0 / 1",   "1 = yes, 0 = no",                              False),
    ("underlying_liver",      "Person",       "0 / 1",   "1 = yes, 0 = no",                              False),
    ("underlying_immunocomp", "Person",       "0 / 1",   "1 = yes, 0 = no",                              False),
    ("underlying_pregnant",   "Person",       "0 / 1",   "1 = yes, 0 = no",                              False),
    ("underlying_disease",    "Person",       "0 / 1",   "1 = has any underlying, 0 = none",             False),
    ("underlying_list",       "Person",       "string",  "Free text: HT, DM",                             False),

    # Place
    ("province",              "Place",        "string",  "Free text",                                      False),
    ("district",              "Place",        "string",  "Free text",                                      False),
    ("subdistrict",           "Place",        "string",  "Free text",                                      False),
    ("village",               "Place",        "string",  "Free text",                                      False),
    ("address",               "Place",        "string",  "Free text",                                      False),

    # Time
    ("date_onset",            "Time",         "date",    "YYYY-MM-DD — e.g. 2024-03-15",                  True),
    ("time_onset",            "Time",         "time",    "HH:MM — e.g. 14:30",                            False),
    ("date_exposure",         "Time",         "date",    "YYYY-MM-DD",                                     False),
    ("date_report",           "Time",         "date",    "YYYY-MM-DD",                                     False),
    ("date_admitted",         "Time",         "date",    "YYYY-MM-DD",                                     False),
    ("date_discharge",        "Time",         "date",    "YYYY-MM-DD",                                     False),
    ("date_death",            "Time",         "date",    "YYYY-MM-DD",                                     False),

    # Case
    ("case_status",           "Case Status",  "0 / 1",   "1 = case (ill), 0 = control/non-case",          False),
    ("diagnosis",             "Case Status",  "string",  "Free text: Cholera, Norovirus",                 False),

    # Symptoms
    ("symptom_fever",         "Symptoms",     "0 / 1",   "1 = present, 0 = absent",                      False),
    ("symptom_cough",         "Symptoms",     "0 / 1",   "1 = present, 0 = absent",                      False),
    ("symptom_diarrhea",      "Symptoms",     "0 / 1",   "1 = present, 0 = absent",                      False),
    ("symptom_vomiting",      "Symptoms",     "0 / 1",   "1 = present, 0 = absent",                      False),
    ("symptom_nausea",        "Symptoms",     "0 / 1",   "1 = present, 0 = absent",                      False),
    ("symptom_abdominal_pain","Symptoms",     "0 / 1",   "1 = present, 0 = absent",                      False),
    ("symptom_headache",      "Symptoms",     "0 / 1",   "1 = present, 0 = absent",                      False),
    ("symptom_myalgia",       "Symptoms",     "0 / 1",   "1 = present, 0 = absent",                      False),
    ("symptom_rash",          "Symptoms",     "0 / 1",   "1 = present, 0 = absent",                      False),
    ("symptom_dyspnea",       "Symptoms",     "0 / 1",   "1 = present, 0 = absent",                      False),
    ("symptom_chest_pain",    "Symptoms",     "0 / 1",   "1 = present, 0 = absent",                      False),
    ("symptom_jaundice",      "Symptoms",     "0 / 1",   "1 = present, 0 = absent",                      False),
    ("symptom_fatigue",       "Symptoms",     "0 / 1",   "1 = present, 0 = absent",                      False),

    # Exposure
    ("exposure_water",        "Exposure",     "0 / 1",   "1 = exposed, 0 = not",                          False),
    ("exposure_food_rice",    "Exposure",     "0 / 1",   "1 = exposed, 0 = not",                          False),
    ("exposure_food_chicken", "Exposure",     "0 / 1",   "1 = exposed, 0 = not",                          False),
    ("exposure_food_pork",    "Exposure",     "0 / 1",   "1 = exposed, 0 = not",                          False),
    ("exposure_food_beef",    "Exposure",     "0 / 1",   "1 = exposed, 0 = not",                          False),
    ("exposure_food_seafood", "Exposure",     "0 / 1",   "1 = exposed, 0 = not",                          False),
    ("exposure_food_egg",     "Exposure",     "0 / 1",   "1 = exposed, 0 = not",                          False),
    ("exposure_food_salad",   "Exposure",     "0 / 1",   "1 = exposed, 0 = not",                          False),
    ("exposure_food_dessert", "Exposure",     "0 / 1",   "1 = exposed, 0 = not",                          False),
    ("exposure_animal_contact","Exposure",    "0 / 1",   "1 = exposed, 0 = not",                          False),
    ("exposure_crowded_place","Exposure",     "0 / 1",   "1 = exposed, 0 = not",                          False),
    ("exposure_shared_toilet","Exposure",     "0 / 1",   "1 = exposed, 0 = not",                          False),
    ("exposure_travel",       "Exposure",     "0 / 1",   "1 = exposed, 0 = not",                          False),

    # Outcome
    ("outcome",               "Outcome",      "string",  "Alive / Dead / Unknown",                        False),
    ("hospitalized",          "Outcome",      "0 / 1",   "1 = yes, 0 = no",                              False),
    ("icu",                   "Outcome",      "0 / 1",   "1 = yes, 0 = no",                              False),
    ("self_remit",            "Outcome",      "0 / 1",   "1 = yes, 0 = no",                              False),

    # Lab
    ("lab_confirmed",         "Laboratory",   "0 / 1",   "1 = confirmed, 0 = not",                       False),
    ("lab_result",            "Laboratory",   "string",  "positive / negative / pending / unknown",       False),
    ("lab_specimen",          "Laboratory",   "string",  "Free text: stool, blood, nasal swab",           False),

    # Treatment
    ("treatment_self",        "Treatment",    "0 / 1",   "1 = yes, 0 = no",                              False),
    ("treatment_drugstore",   "Treatment",    "0 / 1",   "1 = yes, 0 = no",                              False),
    ("treatment_opd",         "Treatment",    "0 / 1",   "1 = yes, 0 = no",                              False),
    ("treatment_ipd",         "Treatment",    "0 / 1",   "1 = yes, 0 = no",                              False),
]

# ── Section colors ─────────────────────────────────────────────────────────────
_SEC_COLORS: dict[str, tuple[str, str]] = {
    "Identifier":  ("0f3460", "e2e8f0"),
    "Person":      ("1e3a5f", "93c5fd"),
    "Place":       ("14432a", "86efac"),
    "Time":        ("3b2f0a", "fcd34d"),
    "Case Status": ("3b0f0f", "fca5a5"),
    "Symptoms":    ("2e1065", "c4b5fd"),
    "Exposure":    ("422006", "fdba74"),
    "Outcome":     ("3b0f0f", "fca5a5"),
    "Laboratory":  ("0c3038", "67e8f9"),
    "Treatment":   ("1e3a5f", "a5b4fc"),
}

# ── Sample data rows ───────────────────────────────────────────────────────────
_SAMPLES = [
    {
        "case_id": "PT-001", "age": 32, "sex": "Male", "occupation": "Farmer",
        "nationality": "Thai", "underlying_ht": 0, "underlying_dm": 0,
        "underlying_copd": 0, "underlying_cad": 0, "underlying_ckd": 0,
        "underlying_liver": 0, "underlying_immunocomp": 0, "underlying_pregnant": 0,
        "underlying_disease": 0, "underlying_list": "",
        "province": "Chiang Mai", "district": "Mueang", "subdistrict": "Suthep",
        "village": "Moo 3", "address": "12 Moo 3 Suthep",
        "date_onset": "2024-03-10", "time_onset": "08:00", "date_exposure": "2024-03-08",
        "date_report": "2024-03-11", "date_admitted": "2024-03-11",
        "date_discharge": "2024-03-13", "date_death": "",
        "case_status": 1, "diagnosis": "Norovirus gastroenteritis",
        "symptom_fever": 1, "symptom_cough": 0, "symptom_diarrhea": 1,
        "symptom_vomiting": 1, "symptom_nausea": 1, "symptom_abdominal_pain": 1,
        "symptom_headache": 0, "symptom_myalgia": 0, "symptom_rash": 0,
        "symptom_dyspnea": 0, "symptom_chest_pain": 0, "symptom_jaundice": 0,
        "symptom_fatigue": 1,
        "exposure_water": 0, "exposure_food_rice": 1, "exposure_food_chicken": 0,
        "exposure_food_pork": 0, "exposure_food_beef": 0, "exposure_food_seafood": 1,
        "exposure_food_egg": 0, "exposure_food_salad": 1, "exposure_food_dessert": 0,
        "exposure_animal_contact": 0, "exposure_crowded_place": 1,
        "exposure_shared_toilet": 0, "exposure_travel": 0,
        "outcome": "Alive", "hospitalized": 1, "icu": 0, "self_remit": 0,
        "lab_confirmed": 1, "lab_result": "positive", "lab_specimen": "stool",
        "treatment_self": 0, "treatment_drugstore": 0, "treatment_opd": 0, "treatment_ipd": 1,
    },
    {
        "case_id": "PT-002", "age": 45, "sex": "Female", "occupation": "Teacher",
        "nationality": "Thai", "underlying_ht": 1, "underlying_dm": 0,
        "underlying_copd": 0, "underlying_cad": 0, "underlying_ckd": 0,
        "underlying_liver": 0, "underlying_immunocomp": 0, "underlying_pregnant": 0,
        "underlying_disease": 1, "underlying_list": "HT",
        "province": "Chiang Mai", "district": "Mueang", "subdistrict": "Suthep",
        "village": "Moo 5", "address": "45 Moo 5 Suthep",
        "date_onset": "2024-03-10", "time_onset": "14:00", "date_exposure": "2024-03-08",
        "date_report": "2024-03-11", "date_admitted": "", "date_discharge": "", "date_death": "",
        "case_status": 1, "diagnosis": "Norovirus gastroenteritis",
        "symptom_fever": 0, "symptom_cough": 0, "symptom_diarrhea": 1,
        "symptom_vomiting": 1, "symptom_nausea": 1, "symptom_abdominal_pain": 1,
        "symptom_headache": 1, "symptom_myalgia": 0, "symptom_rash": 0,
        "symptom_dyspnea": 0, "symptom_chest_pain": 0, "symptom_jaundice": 0,
        "symptom_fatigue": 0,
        "exposure_water": 0, "exposure_food_rice": 1, "exposure_food_chicken": 0,
        "exposure_food_pork": 0, "exposure_food_beef": 0, "exposure_food_seafood": 1,
        "exposure_food_egg": 0, "exposure_food_salad": 0, "exposure_food_dessert": 0,
        "exposure_animal_contact": 0, "exposure_crowded_place": 1,
        "exposure_shared_toilet": 0, "exposure_travel": 0,
        "outcome": "Alive", "hospitalized": 0, "icu": 0, "self_remit": 0,
        "lab_confirmed": 0, "lab_result": "pending", "lab_specimen": "",
        "treatment_self": 1, "treatment_drugstore": 1, "treatment_opd": 0, "treatment_ipd": 0,
    },
    {
        "case_id": "PT-003", "age": 28, "sex": "Male", "occupation": "Student",
        "nationality": "Thai", "underlying_ht": 0, "underlying_dm": 0,
        "underlying_copd": 0, "underlying_cad": 0, "underlying_ckd": 0,
        "underlying_liver": 0, "underlying_immunocomp": 0, "underlying_pregnant": 0,
        "underlying_disease": 0, "underlying_list": "",
        "province": "Chiang Mai", "district": "Mueang", "subdistrict": "Chang Phueak",
        "village": "Moo 1", "address": "8 Moo 1 Chang Phueak",
        "date_onset": "2024-03-11", "time_onset": "06:30", "date_exposure": "2024-03-08",
        "date_report": "2024-03-12", "date_admitted": "", "date_discharge": "", "date_death": "",
        "case_status": 0, "diagnosis": "",
        "symptom_fever": 0, "symptom_cough": 0, "symptom_diarrhea": 0,
        "symptom_vomiting": 0, "symptom_nausea": 0, "symptom_abdominal_pain": 0,
        "symptom_headache": 0, "symptom_myalgia": 0, "symptom_rash": 0,
        "symptom_dyspnea": 0, "symptom_chest_pain": 0, "symptom_jaundice": 0,
        "symptom_fatigue": 0,
        "exposure_water": 0, "exposure_food_rice": 1, "exposure_food_chicken": 0,
        "exposure_food_pork": 0, "exposure_food_beef": 0, "exposure_food_seafood": 0,
        "exposure_food_egg": 0, "exposure_food_salad": 0, "exposure_food_dessert": 0,
        "exposure_animal_contact": 0, "exposure_crowded_place": 1,
        "exposure_shared_toilet": 0, "exposure_travel": 0,
        "outcome": "Alive", "hospitalized": 0, "icu": 0, "self_remit": 0,
        "lab_confirmed": 0, "lab_result": "negative", "lab_specimen": "stool",
        "treatment_self": 0, "treatment_drugstore": 0, "treatment_opd": 0, "treatment_ipd": 0,
    },
]


def _make_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _make_font(hex_color: str, bold: bool = False, size: int = 10) -> Font:
    return Font(color=hex_color, bold=bold, name="Calibri", size=size)


def _build_workbook() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()

    # ── Sheet 1: Line List ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Line List"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"  # freeze rows 1-4 (title + header + type + hint)

    col_names = [c[0] for c in _COLS]
    n_cols = len(col_names)

    # Row 1 — Title banner
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    title_cell = ws.cell(row=1, column=1,
        value="BananaEpi — Line List Template  |  Delete sample rows 5-7 before uploading  |  Required columns marked ★")
    title_cell.fill = _make_fill("0d1b2a")
    title_cell.font = Font(color="5eead4", bold=True, name="Calibri", size=11)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Row 2 — Section header bands
    current_sec = None
    for ci, (col_name, section, type_lbl, hint, req) in enumerate(_COLS, start=1):
        if section != current_sec:
            current_sec = section
            bg, fg = _SEC_COLORS.get(section, ("1e2a3a", "e2e8f0"))
            # Find span
            sec_start = ci
            sec_end = ci
            for cj in range(ci, n_cols + 1):
                if _COLS[cj - 1][1] == section:
                    sec_end = cj
                else:
                    break
            if sec_start < sec_end:
                ws.merge_cells(start_row=2, start_column=sec_start, end_row=2, end_column=sec_end)
            cell = ws.cell(row=2, column=sec_start, value=section.upper())
            cell.fill = _make_fill(bg)
            cell.font = Font(color=fg, bold=True, name="Calibri", size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell = ws.cell(row=2, column=ci)
            bg, _ = _SEC_COLORS.get(section, ("1e2a3a", "e2e8f0"))
            cell.fill = _make_fill(bg)
    ws.row_dimensions[2].height = 16

    # Row 3 — Column names
    for ci, (col_name, section, type_lbl, hint, req) in enumerate(_COLS, start=1):
        label = f"★ {col_name}" if req else col_name
        cell = ws.cell(row=3, column=ci, value=label)
        bg, _ = _SEC_COLORS.get(section, ("1e2a3a", "e2e8f0"))
        cell.fill = _make_fill(bg)
        cell.font = Font(
            color="f87171" if req else "e2e8f0",
            bold=True, name="Calibri", size=10
        )
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        cell.border = _border
    ws.row_dimensions[3].height = 18

    # Row 4 — Type / hint row
    for ci, (col_name, section, type_lbl, hint, req) in enumerate(_COLS, start=1):
        cell = ws.cell(row=4, column=ci, value=f"{type_lbl}  |  {hint}")
        cell.fill = _make_fill("0d1b2a")
        cell.font = Font(color="64748b", italic=True, name="Calibri", size=8)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        cell.border = _border
    ws.row_dimensions[4].height = 15

    # Rows 5-7 — Sample data
    for ri, sample in enumerate(_SAMPLES, start=5):
        for ci, (col_name, section, type_lbl, hint, req) in enumerate(_COLS, start=1):
            val = sample.get(col_name, "")
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = _make_fill("111f2e" if ri % 2 == 1 else "0d1825")
            cell.font = Font(color="cbd5e1", name="Calibri", size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = _border
        ws.row_dimensions[ri].height = 16

    # Column widths
    for ci, (col_name, *_) in enumerate(_COLS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = max(len(col_name) + 2, 14)

    # ── Sheet 2: Instructions ──────────────────────────────────────────────────
    wi = wb.create_sheet("Instructions")
    wi.sheet_view.showGridLines = False
    wi.column_dimensions["A"].width = 5
    wi.column_dimensions["B"].width = 22
    wi.column_dimensions["C"].width = 60

    def irow(row, col, val, bold=False, color="e2e8f0", size=10, bg=None):
        c = wi.cell(row=row, column=col, value=val)
        c.font = Font(color=color, bold=bold, name="Calibri", size=size)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        if bg:
            c.fill = _make_fill(bg)
        return c

    r = 1
    wi.merge_cells(f"B{r}:C{r}")
    irow(r, 2, "BananaEpi — Line List Instructions", bold=True, color="5eead4", size=14, bg="0d1b2a")
    wi.row_dimensions[r].height = 28
    r += 1

    steps = [
        ("Step 1", "Download this template",
         "This file already has all supported columns. Delete columns you don't need. Keep at least: case_id, age, sex, date_onset."),
        ("Step 2", "Fill in your data",
         "Replace the 3 sample rows (rows 5-7) with your real data. Delete rows 5-7 before uploading. Do NOT change the column header names in row 3."),
        ("Step 3", "Date format",
         "All dates must be YYYY-MM-DD (e.g. 2024-03-15). Leave date cells empty for controls / non-cases — do not fill with 0 or '-'."),
        ("Step 4", "Boolean columns (0 / 1)",
         "All symptom_*, exposure_*, underlying_*, hospitalized, lab_confirmed columns accept only 1 (yes/present/exposed) or 0 (no/absent/not exposed). Leave empty if unknown."),
        ("Step 5", "Sex column",
         "Use exactly: Male  Female  Unknown  (capital first letter). Other values become Unknown."),
        ("Step 6", "Outcome column",
         "Use exactly: Alive  Dead  Unknown"),
        ("Step 7", "Lab result column",
         "Use exactly: positive  negative  pending  unknown  (all lowercase)"),
        ("Step 8", "Custom symptoms",
         "Add any column starting with symptom_  (e.g. symptom_bleeding) and it will be auto-detected and included in symptom analysis."),
        ("Step 9", "Custom exposures",
         "Add any column starting with exposure_  (e.g. exposure_well_water) and it will be auto-detected and included in risk ratio analysis."),
        ("Step 10", "case_status (case-control)",
         "If you have both cases and controls in one file: 1 = case (ill person), 0 = control. Leave empty if all rows are cases."),
        ("Step 11", "Missing values",
         "Leave the cell completely empty. Do NOT type: N/A, -, unknown, null, 0 in date columns. Empty = missing."),
        ("Step 12", "Upload",
         "Save as .xlsx or .csv (UTF-8). Go to BananaEpi homepage → drag and drop your file → follow the steps."),
    ]

    for step, title, desc in steps:
        r += 1
        irow(r, 2, step, bold=True, color="5eead4", size=10, bg="0f2027")
        irow(r, 3, title, bold=True, color="f1f5f9", size=10, bg="0f2027")
        wi.row_dimensions[r].height = 18
        r += 1
        irow(r, 3, desc, color="94a3b8", size=9, bg="0a1520")
        wi.row_dimensions[r].height = 30
        r += 1  # spacer

    return wb


@router.get("")
def download_template():
    wb = _build_workbook()
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="epiassist_template.xlsx"'},
    )
